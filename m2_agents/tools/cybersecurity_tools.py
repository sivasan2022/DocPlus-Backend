from __future__ import annotations

import json
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from backend.graph.schema import SourceType
from m2_agents.core.trace_ai import traced_tool


NVD_CVE_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"
DEFAULT_SBOM_PATH = Path("data/sbom/Pulse_Oximeter_SBOM_Report.xlsx")
DEFAULT_CACHE_DIR = Path("data/runtime")
DEFAULT_RESULTS_PER_PAGE = 20
DEFAULT_MAX_CVES_PER_COMPONENT = 5
DEFAULT_PUBLIC_DELAY_SECONDS = 6.2


def resolve_sbom_path(path: str | None = None) -> Path:
    candidates = [
        Path(path) if path else None,
        Path(os.getenv("MEDTRACE_SBOM_PATH", "")) if os.getenv("MEDTRACE_SBOM_PATH") else None,
        DEFAULT_SBOM_PATH,
    ]
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    checked = ", ".join(str(item) for item in candidates if item)
    raise FileNotFoundError(f"No SBOM workbook found. Checked: {checked}")


def load_sbom_components(path: str | Path | None = None) -> list[dict[str, Any]]:
    workbook_path = resolve_sbom_path(str(path) if path else None)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet_name = "SBOM Report" if "SBOM Report" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header(value) for value in rows[0]]
    components: list[dict[str, Any]] = []
    for index, row in enumerate(rows[1:], start=1):
        raw = {headers[column]: _cell_value(value) for column, value in enumerate(row) if column < len(headers)}
        component = raw.get("component", "").strip()
        version = raw.get("version", "").strip()
        if not component or not version:
            continue
        cpe = raw.get("cpe_candidate", "").strip()
        purl = raw.get("package_url", "").strip()
        components.append(
            {
                "id": f"SBOM-XLSX-{index:03d}",
                "component": component,
                "version": version,
                "supplier": raw.get("supplier", "").strip(),
                "purpose": raw.get("purpose", "").strip(),
                "cybersecurity_relevance": raw.get("cybersecurity_relevance", "").strip(),
                "purl": purl,
                "cpe": cpe,
                "license": raw.get("license", "").strip(),
                "update_status": raw.get("update_status", "").strip(),
                "exploitability_context": raw.get("exploitability_context", "").strip(),
                "network_exposure": raw.get("network_exposure", "").strip(),
                "compensating_controls": raw.get("compensating_controls", "").strip(),
                "source_type": SourceType.EXTRACTED.value,
                "source_artifact": str(workbook_path),
                "cpe_resolution": "provided_by_sbom" if cpe else "missing_cpe",
                "cpe_match_confidence": "high" if cpe else "none",
            }
        )
    return components


def scan_components_against_nvd(
    components: list[dict[str, Any]],
    *,
    max_components: int | None = None,
    max_cves_per_component: int = DEFAULT_MAX_CVES_PER_COMPONENT,
    delay_seconds: float | None = None,
    timeout_seconds: float = 30.0,
) -> dict[str, Any]:
    started_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    delay = _request_delay(delay_seconds)
    selected = components[:max_components] if max_components else components
    findings: list[dict[str, Any]] = []
    component_results: list[dict[str, Any]] = []
    request_count = 0
    errors: list[dict[str, str]] = []
    with httpx.Client(timeout=timeout_seconds, headers=_headers()) as client:
        for index, component in enumerate(selected):
            if index and delay > 0:
                time.sleep(delay)
            cpe = str(component.get("cpe") or "").strip()
            if not cpe:
                component_results.append(
                    _component_result(component, "not_queried", 0, "No CPE was present in the SBOM.")
                )
                continue
            try:
                cves = query_nvd_for_cpe(
                    client,
                    cpe,
                    max_results=max_cves_per_component,
                    results_per_page=min(DEFAULT_RESULTS_PER_PAGE, max_cves_per_component),
                )
                request_count += cves["request_count"]
            except Exception as exc:
                errors.append({"component": str(component.get("component", "")), "cpe": cpe, "error": str(exc)})
                component_results.append(_component_result(component, "query_error", 0, str(exc)))
                continue
            component_findings = [
                _finding_from_nvd(component, item)
                for item in cves["items"]
                if _cve_status(item).lower() != "rejected"
            ]
            findings.extend(component_findings)
            if component_findings:
                note = f"{len(component_findings)} CVE record(s) returned by NVD for this CPE."
                status = "open_cves_found"
            else:
                note = f"No known CVEs found for this component/version combination as of {started_at[:10]}."
                status = "no_known_cves"
            component_results.append(_component_result(component, status, len(component_findings), note))
    finished_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    severity_counts: dict[str, int] = {}
    for finding in findings:
        severity = str(finding.get("severity") or "UNKNOWN")
        severity_counts[severity] = severity_counts.get(severity, 0) + 1
    return {
        "query_started_at": started_at,
        "query_finished_at": finished_at,
        "source_type": SourceType.EXTRACTED.value,
        "source": "National Vulnerability Database CVE API 2.0",
        "api_url": NVD_CVE_API,
        "rate_limit_policy": {
            "delay_seconds_between_component_queries": delay,
            "note": "Default delay is set for NVD public API use without an API key.",
        },
        "sbom_component_count": len(components),
        "queried_component_count": len(selected),
        "request_count": request_count,
        "finding_count": len(findings),
        "severity_counts": severity_counts,
        "components": component_results,
        "findings": findings,
        "errors": errors,
    }


@traced_tool("NVD CVE API 2.0")
def query_nvd_for_cpe(
    client: httpx.Client,
    cpe: str,
    *,
    max_results: int,
    results_per_page: int = DEFAULT_RESULTS_PER_PAGE,
) -> dict[str, Any]:
    items: list[dict[str, Any]] = []
    start_index = 0
    request_count = 0
    while len(items) < max_results:
        params = {
            "cpeName": cpe,
            "resultsPerPage": min(results_per_page, max_results - len(items)),
            "startIndex": start_index,
        }
        response = client.get(NVD_CVE_API, params=params)
        request_count += 1
        if response.status_code == 429:
            time.sleep(max(DEFAULT_PUBLIC_DELAY_SECONDS * 2, 12.0))
            response = client.get(NVD_CVE_API, params=params)
            request_count += 1
        response.raise_for_status()
        payload = response.json()
        page_items = [item.get("cve", {}) for item in payload.get("vulnerabilities", []) if item.get("cve")]
        items.extend(page_items)
        total = int(payload.get("totalResults") or 0)
        start_index += int(payload.get("resultsPerPage") or len(page_items) or results_per_page)
        if start_index >= total or not page_items:
            break
    return {"items": items[:max_results], "request_count": request_count}


def cache_path(device_id: str) -> Path:
    safe = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in (device_id or "device"))
    return DEFAULT_CACHE_DIR / f"cybersecurity_scan_{safe}.json"


def load_cached_scan(device_id: str) -> dict[str, Any] | None:
    path = cache_path(device_id)
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def save_cached_scan(device_id: str, payload: dict[str, Any]) -> None:
    path = cache_path(device_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _finding_from_nvd(component: dict[str, Any], cve: dict[str, Any]) -> dict[str, Any]:
    cvss = _cvss(cve)
    cve_id = str(cve.get("id") or "")
    severity = cvss.get("severity") or "UNKNOWN"
    score = cvss.get("score")
    status = _cve_status(cve)
    description = _english_description(cve)
    weaknesses = _weakness_ids(cve)
    references = _reference_urls(cve)
    return {
        "component": component.get("component", ""),
        "version": component.get("version", ""),
        "supplier": component.get("supplier", ""),
        "cpe": component.get("cpe", ""),
        "purl": component.get("purl", ""),
        "cve_id": cve_id,
        "severity": severity,
        "cvss_score": score,
        "cvss_version": cvss.get("version", ""),
        "cvss_vector": cvss.get("vector_string", ""),
        "nvd_status": status,
        "open_unpatched": status.lower() != "rejected",
        "patched_status": "Requires review; NVD returned this CPE as affected and no local patch disposition is recorded in the SBOM.",
        "exploitability_note": _exploitability_note(component, cvss, description, weaknesses),
        "description_summary": _description_summary(description),
        "description_excerpt": _short_excerpt(description),
        "weaknesses": weaknesses,
        "references": references,
        "attack_vector": cvss.get("attack_vector") or "UNKNOWN",
        "source_type": SourceType.EXTRACTED.value,
        "source": "NVD CVE API 2.0",
        "source_url": f"https://nvd.nist.gov/vuln/detail/{cve_id}" if cve_id else "https://nvd.nist.gov/vuln",
        "review_status": "security_review_required",
        "controlled_status": "external_authority",
    }


def _component_result(component: dict[str, Any], status: str, cve_count: int, note: str) -> dict[str, Any]:
    return {
        **component,
        "nvd_status": status,
        "cve_count": cve_count,
        "nvd_note": note,
    }


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("/", " ").replace("-", " ").replace(" ", "_")


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _headers() -> dict[str, str]:
    headers = {"User-Agent": "MedTraceAI-Cybersecurity-Agent/1.0"}
    api_key = os.getenv("NVD_API_KEY")
    if api_key:
        headers["apiKey"] = api_key
    return headers


def _request_delay(value: float | None) -> float:
    if value is not None:
        return max(0.0, float(value))
    if os.getenv("NVD_API_KEY"):
        return float(os.getenv("MEDTRACE_NVD_DELAY_SECONDS", "0.7"))
    return float(os.getenv("MEDTRACE_NVD_DELAY_SECONDS", str(DEFAULT_PUBLIC_DELAY_SECONDS)))


def _cvss(cve: dict[str, Any]) -> dict[str, Any]:
    metrics = cve.get("metrics") or {}
    for key in ("cvssMetricV40", "cvssMetricV31", "cvssMetricV30", "cvssMetricV2"):
        values = metrics.get(key) or []
        if not values:
            continue
        metric = values[0]
        data = metric.get("cvssData") or {}
        return {
            "version": str(data.get("version") or key.replace("cvssMetric", "")),
            "score": data.get("baseScore"),
            "severity": data.get("baseSeverity") or metric.get("baseSeverity") or "UNKNOWN",
            "attack_vector": data.get("attackVector") or data.get("accessVector") or "UNKNOWN",
            "vector_string": data.get("vectorString") or "",
        }
    return {"version": "", "score": None, "severity": "UNKNOWN", "attack_vector": "UNKNOWN", "vector_string": ""}


def _english_description(cve: dict[str, Any]) -> str:
    for item in cve.get("descriptions", []) or []:
        if str(item.get("lang", "")).lower() == "en":
            return str(item.get("value") or "")
    return ""


def _weakness_ids(cve: dict[str, Any]) -> list[str]:
    ids: list[str] = []
    for weakness in cve.get("weaknesses", []) or []:
        for item in weakness.get("description", []) or []:
            value = str(item.get("value") or "").strip()
            if value and value not in ids:
                ids.append(value)
    return ids


def _reference_urls(cve: dict[str, Any], limit: int = 5) -> list[str]:
    urls: list[str] = []
    references = cve.get("references", [])
    if isinstance(references, dict):
        items = references.get("referenceData", []) or []
    else:
        items = references or []
    for item in items:
        url = str(item.get("url") or "").strip()
        if url and url not in urls:
            urls.append(url)
        if len(urls) >= limit:
            break
    return urls


def _description_summary(description: str) -> str:
    text = description.lower()
    impacts = []
    if "denial of service" in text or "dos" in text:
        impacts.append("availability impact")
    if "remote" in text:
        impacts.append("remote exploitation context")
    if "buffer" in text or "memory" in text or "overflow" in text:
        impacts.append("memory-safety concern")
    if "certificate" in text or "tls" in text or "ssl" in text:
        impacts.append("secure-communication concern")
    if not impacts:
        impacts.append("security impact described by NVD")
    return "NVD describes " + ", ".join(impacts) + "; review the CVE record for full advisory detail."


def _short_excerpt(description: str, limit: int = 160) -> str:
    text = " ".join(description.split())
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def _exploitability_note(
    component: dict[str, Any],
    cvss: dict[str, Any],
    description: str,
    weaknesses: list[str],
) -> str:
    score = cvss.get("score")
    severity = cvss.get("severity") or "UNKNOWN"
    attack_vector = cvss.get("attack_vector") or "UNKNOWN"
    score_text = f"CVSS {score}" if score is not None else "CVSS score unavailable"
    vulnerability = _description_based_vulnerability_summary(description, weaknesses)
    exposure = _description_based_exposure(component, description, attack_vector)
    controls = _description_based_controls(component, description, weaknesses)
    weakness_text = f" Weakness: {', '.join(weaknesses[:3])}." if weaknesses else ""
    return (
        f"{severity} {score_text}; attack vector {attack_vector}. "
        f"Vulnerability: {vulnerability}{weakness_text} "
        f"Exposure: {exposure} Controls: {controls}"
    )


def _description_based_vulnerability_summary(description: str, weaknesses: list[str]) -> str:
    text = " ".join(str(description or "").split())
    lower = text.lower()
    if not text or len(text) < 30:
        return "Limited public description available for this CVE; review the NVD record before assigning device impact."

    if "only one tls certificate" in lower and "trusted" in lower:
        return (
            "NVD describes incomplete TLS certificate-chain validation, where a connection may be accepted "
            "because one certificate in an otherwise untrusted chain is trusted."
        )
    if "\\0" in text or "\x00" in text or "nul" in lower or "null character" in lower:
        if "common name" in lower or " cn " in f" {lower} ":
            return (
                "NVD describes certificate host-name validation that mishandles embedded NUL characters "
                "in the certificate Common Name, enabling a crafted certificate-name mismatch."
            )
        if "subject alternative name" in lower or "subjectaltnames" in lower:
            return (
                "NVD describes certificate host-name validation that mishandles embedded NUL characters "
                "in Subject Alternative Name data, enabling a crafted certificate-name mismatch."
            )
        return "NVD describes input validation weakness involving embedded NUL characters in certificate or string handling."
    if "denial of service" in lower or " dos " in f" {lower} ":
        return "NVD describes a denial-of-service condition that may let an attacker disrupt availability of the affected software path."
    if any(term in lower for term in ["buffer overflow", "heap overflow", "stack overflow", "out-of-bounds"]):
        return "NVD describes a memory-boundary weakness that may permit crash or code-execution style impact depending on deployment."
    if any(term in lower for term in ["improper authentication", "authentication bypass", "bypass authentication"]):
        return "NVD describes an authentication or access-control bypass condition in the affected component."
    if "certificate" in lower or "tls" in lower or "ssl" in lower or "x.509" in lower:
        return "NVD describes a certificate or TLS validation weakness in the affected secure-communication path."
    if "integer overflow" in lower:
        return "NVD describes integer overflow behavior that may affect memory handling or validation logic."
    if weaknesses:
        return f"NVD classifies the CVE under {', '.join(weaknesses[:2])}; review the public description for deployment-specific impact."
    return "NVD describes a component-specific software vulnerability; review the public description for deployment-specific impact."


def _description_based_exposure(component: dict[str, Any], description: str, attack_vector: str) -> str:
    exposure = component.get("network_exposure") or "Exposure not specified in SBOM."
    lower = str(description or "").lower()
    if "certificate" in lower or "tls" in lower or "ssl" in lower or "x.509" in lower:
        return (
            "Relevant when this component validates TLS peers, update endpoints, cloud links, or other certificate-backed channels. "
            f"SBOM exposure: {exposure}"
        )
    if attack_vector == "NETWORK":
        return f"Network-reachable exploitation should be assessed against the device communication paths. SBOM exposure: {exposure}"
    if attack_vector == "LOCAL":
        return f"Local or physically mediated exploitation should be assessed against service, update, or peripheral workflows. SBOM exposure: {exposure}"
    if attack_vector == "ADJACENT_NETWORK":
        return f"Adjacent-network exploitation should be assessed against Wi-Fi/BLE pairing and local-network workflows. SBOM exposure: {exposure}"
    return f"Assess the CVE against the component's actual deployment and interface reachability. SBOM exposure: {exposure}"


def _description_based_controls(component: dict[str, Any], description: str, weaknesses: list[str]) -> str:
    controls = component.get("compensating_controls") or "No compensating controls listed in SBOM."
    lower = str(description or "").lower()
    weakness_text = " ".join(weaknesses).lower()
    if "only one tls certificate" in lower and "trusted" in lower:
        specific = "verify full certificate-chain validation, require trusted root and hostname match, and regression-test rejected-chain cases"
    elif "\\0" in description or "\x00" in description or "nul" in lower or "null character" in lower:
        specific = "reject embedded NUL characters in certificate names, prefer patched TLS hostname-validation routines, and test crafted certificate names"
    elif "denial of service" in lower or " dos " in f" {lower} ":
        specific = "patch affected code paths, constrain exposed interfaces, and confirm watchdog/recovery behavior for availability loss"
    elif any(term in lower for term in ["buffer overflow", "heap overflow", "stack overflow", "out-of-bounds"]) or "cwe-119" in weakness_text:
        specific = "apply the vendor fix, keep compiler/runtime hardening enabled, and add bounds/fuzz regression coverage"
    elif "certificate" in lower or "tls" in lower or "ssl" in lower or "x.509" in lower:
        specific = "update the TLS library, enforce strict certificate validation, and regression-test invalid certificate cases"
    else:
        specific = "review vendor fix status, confirm exploit preconditions, and document whether compensating controls cover the affected interface"
    return f"{specific}. SBOM controls: {controls}"


def _cve_status(cve: dict[str, Any]) -> str:
    return str(cve.get("vulnStatus") or cve.get("cveTags") or "Analyzed")

from __future__ import annotations

import hashlib
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Iterable

from backend.graph.schema import SourceType
from backend.graph.store import JsonGraphStore


@dataclass
class StructuredImportSummary:
    artifacts: int = 0
    requirements: int = 0
    risks: int = 0
    tests: int = 0
    test_runs: int = 0
    complaints: int = 0
    capas: int = 0
    risk_controls: int = 0
    evidence: int = 0

    def as_dict(self) -> dict[str, int]:
        return {
            "artifacts": self.artifacts,
            "requirements": self.requirements,
            "risks": self.risks,
            "tests": self.tests,
            "test_runs": self.test_runs,
            "complaints": self.complaints,
            "capas": self.capas,
            "risk_controls": self.risk_controls,
            "evidence": self.evidence,
        }


def ingest_structured_m1_artifacts(
    source: str | Path,
    device_id: str,
    current_firmware: str,
    graph_store: JsonGraphStore,
) -> dict[str, int]:
    source_path = Path(source)
    summary = StructuredImportSummary()
    with TemporaryDirectory() as tmp:
        root = _materialize_source(source_path, Path(tmp))
        files = {path.name: path for path in root.rglob("*") if path.is_file()}
        if not files:
            return summary.as_dict()
        metadata = _artifact_metadata(files, root)

        for path in sorted(files.values()):
            _create_artifact_evidence(graph_store, device_id, path, root, summary, metadata)

        _ingest_srs(files.get("SRS_Measurement_Firmware_Requirements_Unique.xlsx"), graph_store, device_id, current_firmware, summary)
        _ingest_alarm_requirements(files.get("REQ_Alarm_Behavior_and_Latency.xlsx"), graph_store, device_id, current_firmware, summary)
        _ingest_display_requirements(files.get("REQ_Display_UI_Behavior_Unique.xlsx"), graph_store, device_id, current_firmware, summary)
        _ingest_hazard_analysis(files.get("Hazard_Analysis_PulseOx.xlsx"), graph_store, device_id, summary)
        _ingest_risk_control_matrix(files.get("Risk_Control_Verification_Matrix.xlsx"), graph_store, device_id, current_firmware, summary)
        _ingest_rtm(files.get("RTM_Master.xlsx"), graph_store, device_id, current_firmware, summary)
        _ingest_release_and_logs(files, graph_store, device_id, current_firmware)

    return summary.as_dict()


def _materialize_source(source_path: Path, tmp: Path) -> Path:
    if source_path.is_dir():
        return source_path
    if source_path.is_file() and source_path.suffix.lower() == ".zip":
        root = tmp / "m1_artifacts"
        with zipfile.ZipFile(source_path) as archive:
            archive.extractall(root)
        return root
    if source_path.is_file():
        root = tmp / "single"
        root.mkdir(parents=True, exist_ok=True)
        target = root / source_path.name
        target.write_bytes(source_path.read_bytes())
        return root
    raise FileNotFoundError(f"Source not found: {source_path}")


def _create_artifact_evidence(
    store: JsonGraphStore,
    device_id: str,
    path: Path,
    root: Path,
    summary: StructuredImportSummary,
    metadata_by_key: dict[str, dict[str, Any]],
) -> tuple[str, str]:
    rel_path = str(path.relative_to(root)).replace("\\", "/")
    metadata = _metadata_for(path, rel_path, metadata_by_key)
    stem = _slug(path.stem)
    digest = _file_digest(path)
    doc_id = f"DOC-{stem}"
    evidence_id = f"EVID-{stem}"
    existed = store.get_node(evidence_id) is not None
    artifact_type = metadata.get("artifact_class") or _artifact_type(path.name)
    controlled_status = _string(metadata.get("controlled_status")) or ("approved" if path.suffix.lower() in {".xlsx", ".pdf"} else "raw")
    review_status = _string(metadata.get("review_status")) or ("approved" if controlled_status == "approved" else "needs_review")
    objective = _bool(metadata.get("objective_evidence"), _objective_default(path.name, artifact_type))
    firmware_version = _infer_firmware_from_text(path.name) or _infer_firmware_from_text(rel_path)
    metadata_source_type = _string(metadata.get("source_type")) or SourceType.CONTROLLED.value
    store.upsert_node(
        doc_id,
        ["SourceDocument"],
        title=path.stem.replace("_", " "),
        file_name=path.name,
        source_path=rel_path,
        extension=path.suffix.lower(),
        controlled_status=controlled_status,
        review_status=review_status,
        source_type=metadata_source_type,
        content_hash=digest,
        revision=metadata.get("revision") or metadata.get("revision_or_publication_date") or "M1-2026-07-01",
        owner=metadata.get("owner"),
        objective_evidence=objective,
    )
    store.upsert_node(
        evidence_id,
        ["Evidence", "EvidenceArtifact"],
        title=path.stem.replace("_", " "),
        file_name=path.name,
        artifact_type=artifact_type,
        doc_type=artifact_type,
        source_path=rel_path,
        revision="M1-2026-07-01",
        hash=digest,
        controlled_status=controlled_status,
        review_status=review_status,
        source_type=metadata_source_type,
        approval_status=metadata.get("approval_status") or "approved for M1 demo evidence",
        objective_evidence=objective,
        firmware_version=firmware_version,
        page_number=1,
    )
    store.upsert_edge(device_id, evidence_id, "HAS_DOCUMENT", rationale="DocPlus+ M1 source pack artifact")
    store.upsert_edge(evidence_id, doc_id, "STORED_AS", rationale="Evidence artifact physical source document")
    if not existed:
        summary.evidence += 1
    summary.artifacts += 1
    return evidence_id, doc_id


def _ingest_srs(
    path: Path | None,
    store: JsonGraphStore,
    device_id: str,
    current_firmware: str,
    summary: StructuredImportSummary,
) -> None:
    for row in _xlsx_rows(path):
        req_id = _clean_id(row.get("requirement id"))
        if not req_id:
            continue
        component_id = _clean_id(row.get("component"))
        risk_id = _clean_id(row.get("risk id"))
        test_id = _clean_id(row.get("verification test"))
        complaint_id = _clean_id(row.get("trace complaints"))
        _upsert_requirement(
            store,
            device_id,
            req_id,
            text=row.get("requirement statement"),
            acceptance=row.get("acceptance criteria"),
            category="Measurement Firmware",
            standard="IEC 62304",
            module=row.get("module"),
            component_id=component_id,
            risk_id=risk_id,
            source_artifact=path.name if path else None,
            summary=summary,
        )
        if component_id:
            _upsert_component(store, device_id, component_id, row.get("component"), row.get("module"))
            store.upsert_edge(component_id, req_id, "AFFECTS", rationale="SRS component-to-requirement trace")
        if risk_id:
            _upsert_risk(store, device_id, risk_id, hazard=row.get("complaint scenario"), summary=summary)
            _link_requirement_risk(store, req_id, risk_id, "SRS risk trace")
        if test_id:
            _upsert_test_case(store, test_id, req_id, current_firmware, row.get("acceptance criteria"), path.name if path else None, summary)
        if complaint_id:
            _upsert_complaint(store, device_id, complaint_id, row.get("complaint scenario"), summary)
            if risk_id:
                store.upsert_edge(complaint_id, risk_id, "INDICATES", rationale="Complaint scenario indicates risk")


def _ingest_alarm_requirements(
    path: Path | None,
    store: JsonGraphStore,
    device_id: str,
    current_firmware: str,
    summary: StructuredImportSummary,
) -> None:
    component_id = "COMP-ALARM-MANAGER"
    _upsert_component(store, device_id, component_id, "alarm_manager", "Alarm Notification")
    for row in _xlsx_rows(path):
        req_id = _clean_id(row.get("requirement id"))
        if not req_id:
            continue
        test_id = _clean_id(row.get("verification test"))
        complaint_id = _clean_id(row.get("complaint link"))
        _upsert_requirement(
            store,
            device_id,
            req_id,
            text=row.get("latency/behavior requirement"),
            acceptance=row.get("acceptance"),
            category="Alarm Behavior and Latency",
            standard="IEC 60601-1-8",
            module="alarm_manager",
            component_id=component_id,
            source_artifact=path.name if path else None,
            alarm_scenario=row.get("alarm scenario"),
            regulatory_note=row.get("iec 60601-1-8 consideration"),
            summary=summary,
        )
        store.upsert_edge(component_id, req_id, "AFFECTS", rationale="Alarm manager requirement trace")
        if test_id:
            _upsert_test_case(store, test_id, req_id, current_firmware, row.get("acceptance"), path.name if path else None, summary)
        if complaint_id:
            _upsert_complaint(store, device_id, complaint_id, row.get("alarm scenario"), summary)


def _ingest_display_requirements(
    path: Path | None,
    store: JsonGraphStore,
    device_id: str,
    current_firmware: str,
    summary: StructuredImportSummary,
) -> None:
    component_id = "COMP-DISPLAY-UI"
    _upsert_component(store, device_id, component_id, "display_task", "Display and UI")
    for row in _xlsx_rows(path):
        req_id = _clean_id(row.get("requirement id"))
        if not req_id:
            continue
        test_id = _clean_id(row.get("verification test"))
        complaint_id = _clean_id(row.get("complaint link"))
        _upsert_requirement(
            store,
            device_id,
            req_id,
            text=row.get("ui requirement"),
            acceptance=row.get("acceptance"),
            category="Display/UI Behavior",
            standard="IEC 62366-1",
            module="display_task",
            component_id=component_id,
            source_artifact=path.name if path else None,
            display_scenario=row.get("display/ui scenario"),
            regulatory_note=row.get("usability/risk note"),
            summary=summary,
        )
        store.upsert_edge(component_id, req_id, "AFFECTS", rationale="Display task requirement trace")
        if test_id:
            _upsert_test_case(store, test_id, req_id, current_firmware, row.get("acceptance"), path.name if path else None, summary)
        if complaint_id:
            _upsert_complaint(store, device_id, complaint_id, row.get("display/ui scenario"), summary)


def _ingest_hazard_analysis(path: Path | None, store: JsonGraphStore, device_id: str, summary: StructuredImportSummary) -> None:
    for row in _xlsx_rows(path):
        risk_id = _clean_id(row.get("risk id"))
        if not risk_id:
            continue
        _upsert_risk(
            store,
            device_id,
            risk_id,
            hazard=row.get("hazardous situation"),
            sequence_of_events=row.get("sequence of events"),
            harm=row.get("potential harm"),
            severity=row.get("severity"),
            probability=row.get("occurrence initial"),
            initial_rpn=row.get("initial rpn"),
            residual_probability=row.get("residual occurrence"),
            residual_rpn=row.get("residual rpn"),
            risk_controls=row.get("risk controls"),
            source_artifact=path.name if path else None,
            summary=summary,
        )
        complaint_id = _clean_id(row.get("complaint link"))
        if complaint_id:
            _upsert_complaint(store, device_id, complaint_id, row.get("hazardous situation"), summary)
            store.upsert_edge(complaint_id, risk_id, "INDICATES", rationale="Hazard analysis complaint linkage")


def _ingest_risk_control_matrix(
    path: Path | None,
    store: JsonGraphStore,
    device_id: str,
    current_firmware: str,
    summary: StructuredImportSummary,
) -> None:
    for row in _xlsx_rows(path):
        risk_id = _clean_id(row.get("risk id"))
        if not risk_id:
            continue
        control_id = f"RC-{risk_id}"
        existed = store.get_node(control_id) is not None
        store.upsert_node(
            control_id,
            ["RiskControl"],
            risk_control=row.get("risk control"),
            verification_method=row.get("verification method"),
            effectiveness_evidence=row.get("effectiveness evidence"),
            residual_risk_decision=row.get("residual risk decision"),
            root_cause_category=row.get("root cause category"),
            source_artifact=path.name if path else None,
        )
        if not existed:
            summary.risk_controls += 1
        _upsert_risk(store, device_id, risk_id, summary=summary)
        store.upsert_edge(risk_id, control_id, "CONTROLLED_BY", rationale="Risk control verification matrix")
        test_id = _clean_id(row.get("verification method"))
        if test_id:
            _upsert_test_case(store, test_id, None, current_firmware, row.get("effectiveness evidence"), path.name if path else None, summary)
            store.upsert_edge(control_id, test_id, "VERIFIED_BY", rationale="Risk control verification method")
        for req_id in _extract_requirement_ids(row.get("risk control")):
            _upsert_requirement(store, device_id, req_id, category="Risk Control", standard="ISO 14971", summary=summary)
            _link_requirement_risk(store, req_id, risk_id, "Risk control matrix trace")
        complaint_id = _clean_id(row.get("complaint link"))
        if complaint_id:
            _upsert_complaint(store, device_id, complaint_id, row.get("root cause category"), summary)
            store.upsert_edge(complaint_id, risk_id, "INDICATES", rationale="Risk control complaint linkage")


def _ingest_rtm(
    path: Path | None,
    store: JsonGraphStore,
    device_id: str,
    current_firmware: str,
    summary: StructuredImportSummary,
) -> None:
    report_evidence = _evidence_id("TEST_REPORT_Complaint100_Regression_FW-v3.4")
    change_id = f"FWCHANGE-{current_firmware}"
    for row in _xlsx_rows(path):
        complaint_id = _clean_id(row.get("complaint"))
        req_id = _clean_id(row.get("requirement"))
        risk_id = _clean_id(row.get("risk"))
        test_id = _clean_id(row.get("verification test"))
        tested_firmware = _infer_firmware_from_text(_row_text(row)) or current_firmware
        test_result = "Pass" if "closed" in _string(row.get("closure status")).lower() else "Needs Review"
        if complaint_id:
            _upsert_complaint(
                store,
                device_id,
                complaint_id,
                row.get("scenario"),
                root_cause=row.get("root cause"),
                closure_status=row.get("closure status"),
                firmware_at_event=tested_firmware,
                investigation_scope=row.get("evidence"),
                summary=summary,
            )
            store.upsert_node(
                change_id,
                ["FirmwareChange"],
                version=current_firmware,
                summary="Firmware change package referenced by complaint traceability records.",
                source_artifact="Firmware_Release_Notes_FW-v3.4.pdf",
            )
            store.upsert_edge(complaint_id, change_id, "OCCURRED_AFTER", rationale="Complaint trace assessed against current firmware change package")
        if req_id:
            _upsert_requirement(store, device_id, req_id, text=row.get("scenario"), category="RTM Master", source_artifact=path.name if path else None, summary=summary)
        if risk_id:
            _upsert_risk(store, device_id, risk_id, hazard=row.get("scenario"), root_cause=row.get("root cause"), summary=summary)
        if req_id and risk_id:
            _link_requirement_risk(store, req_id, risk_id, "RTM master trace")
        if test_id:
            _upsert_test_case(
                store,
                test_id,
                req_id,
                tested_firmware,
                row.get("evidence"),
                path.name if path else None,
                summary,
                test_objective=row.get("scenario"),
                result_text=row.get("evidence"),
            )
            run_id = f"RUN-{test_id}-{_slug(tested_firmware)}"
            _upsert_test_run(
                store,
                run_id,
                test_id,
                tested_firmware,
                report_evidence,
                result=test_result,
                complaint_id=complaint_id,
                summary=summary,
                requirement_id=req_id,
                evidence_summary=row.get("evidence"),
                root_cause=row.get("root cause"),
            )
        if complaint_id and risk_id:
            capa_id = f"CAPA-{complaint_id}"
            existed = store.get_node(capa_id) is not None
            store.upsert_node(
                capa_id,
                ["CAPA"],
                title=f"CAPA decision for {complaint_id}",
                action=row.get("capa decision"),
                root_cause=row.get("root cause"),
                status="Closed" if "closed" in str(row.get("closure status", "")).lower() else "Open",
                source_artifact=path.name if path else None,
            )
            if not existed:
                summary.capas += 1
            store.upsert_edge(complaint_id, capa_id, "TRIGGERS", rationale="RTM complaint CAPA decision")
            store.upsert_edge(capa_id, risk_id, "ADDRESSES", rationale="CAPA addresses traced risk")
            if store.get_node(report_evidence):
                store.upsert_edge(capa_id, report_evidence, "SUPPORTED_BY", rationale="CAPA decision evidence")


def _ingest_release_and_logs(files: dict[str, Path], store: JsonGraphStore, device_id: str, current_firmware: str) -> None:
    fw_id = f"FW-{current_firmware}"
    release = files.get("Firmware_Release_Notes_FW-v3.4.pdf")
    if release:
        change_id = f"FWCHANGE-{current_firmware}"
        store.upsert_node(
            change_id,
            ["FirmwareChange"],
            version=current_firmware,
            release_date="2026-07-01",
            changed_modules="alarm_manager, display_task, power_manager, measurement firmware",
            source_artifact=release.name,
            summary="FW-v3.4 complaint evidence refresh and regression release package.",
        )
        store.upsert_edge(fw_id, change_id, "HAS_CHANGE", rationale="Firmware release notes")
        store.upsert_edge(change_id, _evidence_id(release.stem), "SUPPORTED_BY", rationale="Release notes objective evidence")

    for name in [
        "RAWLOG_Alarm_Event_Queue_FW-v3.4.log.txt",
        "RAWLOG_Display_Task_FW-v3.4.log.txt",
        "RAWLOG_Complaint100_System_FW-v3.4.log.txt",
    ]:
        path = files.get(name)
        if not path:
            continue
        log_id = f"LOG-{_slug(path.stem)}"
        store.upsert_node(
            log_id,
            ["TelemetryLog"],
            file_name=path.name,
            firmware_version=current_firmware,
            log_type=_artifact_type(path.name),
            source_artifact=path.name,
        )
        evidence_id = _evidence_id(path.stem)
        store.upsert_edge(device_id, log_id, "HAS_LOG", rationale="Runtime telemetry source pack")
        store.upsert_edge(log_id, evidence_id, "SUPPORTED_BY", rationale="Telemetry log stored as evidence")


def _upsert_requirement(
    store: JsonGraphStore,
    device_id: str,
    req_id: str,
    text: Any = None,
    acceptance: Any = None,
    category: str | None = None,
    standard: str | None = None,
    module: Any = None,
    component_id: str | None = None,
    risk_id: str | None = None,
    source_artifact: str | None = None,
    summary: StructuredImportSummary | None = None,
    **extra: Any,
) -> None:
    existed = store.get_node(req_id) is not None
    labels = ["Requirement"]
    if category and "firmware" in category.lower():
        labels.append("SoftwareRequirement")
    store.upsert_node(
        req_id,
        labels,
        text=_string(text) or None,
        acceptance_criteria=_string(acceptance) or None,
        category=category,
        standard=standard,
        module=_string(module) or None,
        component_id=component_id,
        risk_id=risk_id,
        status="Active",
        source_artifact=source_artifact,
        **{key: _string(value) for key, value in extra.items() if value not in (None, "")},
    )
    store.upsert_edge(device_id, req_id, "CONTAINS", rationale="DocPlus+ M1 controlled requirement")
    if source_artifact:
        evidence_id = _evidence_id(Path(source_artifact).stem)
        if store.get_node(evidence_id):
            store.upsert_edge(req_id, evidence_id, "SUPPORTED_BY", rationale="Requirement source artifact")
    if summary and not existed:
        summary.requirements += 1


def _upsert_component(store: JsonGraphStore, device_id: str, component_id: str, name: Any, module: Any = None) -> None:
    store.upsert_node(
        component_id,
        ["Component"],
        name=_string(name) or component_id,
        module=_string(module) or None,
        part_type="Software/Firmware" if str(component_id).startswith("COMP-") else "Component",
        safety_relevance="Safety relevant",
    )
    store.upsert_edge(device_id, component_id, "HAS_COMPONENT", rationale="SRS component map")


def _upsert_risk(
    store: JsonGraphStore,
    device_id: str,
    risk_id: str,
    summary: StructuredImportSummary | None = None,
    **properties: Any,
) -> None:
    existed = store.get_node(risk_id) is not None
    store.upsert_node(
        risk_id,
        ["Risk"],
        **{key: _string(value) for key, value in properties.items() if value not in (None, "")},
    )
    store.upsert_edge(device_id, risk_id, "HAS_RISK", rationale="DocPlus+ M1 risk management file")
    source_artifact = properties.get("source_artifact")
    if source_artifact:
        evidence_id = _evidence_id(Path(str(source_artifact)).stem)
        if store.get_node(evidence_id):
            store.upsert_edge(risk_id, evidence_id, "SUPPORTED_BY", rationale="Risk source artifact")
    if summary and not existed:
        summary.risks += 1


def _upsert_test_case(
    store: JsonGraphStore,
    test_id: str,
    req_id: str | None,
    current_firmware: str,
    acceptance: Any,
    source_artifact: str | None,
    summary: StructuredImportSummary,
    **extra: Any,
) -> None:
    existed = store.get_node(test_id) is not None
    store.upsert_node(
        test_id,
        ["Test", "TestCase"],
        name=f"Verification test {test_id}",
        method="Requirements-based verification",
        result="Pass",
        firmware_tested=current_firmware,
        acceptance_criteria=_string(acceptance) or None,
        acceptance_criteria_met=True,
        source_artifact=source_artifact,
        **{key: _string(value) for key, value in extra.items() if value not in (None, "")},
    )
    if req_id:
        store.upsert_edge(req_id, test_id, "VERIFIED_BY", rationale="Controlled traceability matrix")
    if source_artifact:
        evidence_id = _evidence_id(Path(source_artifact).stem)
        if store.get_node(evidence_id):
            store.upsert_edge(test_id, evidence_id, "SUPPORTED_BY", rationale="Test case source artifact")
    if not existed:
        summary.tests += 1


def _upsert_test_run(
    store: JsonGraphStore,
    run_id: str,
    test_id: str,
    current_firmware: str,
    evidence_id: str,
    result: str,
    complaint_id: str | None,
    summary: StructuredImportSummary,
    **extra: Any,
) -> None:
    fw_id = f"FW-{current_firmware}"
    existed = store.get_node(run_id) is not None
    if store.get_node(fw_id) is None:
        store.upsert_node(
            fw_id,
            ["SoftwareVersion", "FirmwareVersion"],
            version=current_firmware,
            release_date="2026-06-30",
            change_summary="Firmware version referenced by controlled test-run evidence.",
            source_type=SourceType.EXTRACTED.value,
        )
    store.upsert_node(
        run_id,
        ["TestRun"],
        test_case_id=test_id,
        firmware_version=current_firmware,
        result=result,
        test_result=result,
        execution_status="Executed",
        complaint_id=complaint_id,
        acceptance_criteria_met=result.lower() == "pass",
        execution_date="2026-06-30",
        source_type=SourceType.CONTROLLED.value,
        controlled_status="approved",
        review_status="approved",
        objective_evidence=True,
        evidence_artifact_id=evidence_id,
        **{key: _string(value) for key, value in extra.items() if value not in (None, "")},
    )
    store.upsert_edge(test_id, run_id, "EXECUTED_AS", rationale="Test case executed as controlled test run")
    store.upsert_edge(run_id, fw_id, "TESTED_ON", rationale="Test run firmware version")
    if store.get_node(evidence_id):
        store.upsert_edge(run_id, evidence_id, "PRODUCED", rationale="Objective regression evidence")
        store.upsert_edge(test_id, evidence_id, "SUPPORTED_BY", rationale="Regression report evidence")
    if not existed:
        summary.test_runs += 1


def _upsert_complaint(
    store: JsonGraphStore,
    device_id: str,
    complaint_id: str,
    description: Any,
    summary: StructuredImportSummary,
    **properties: Any,
) -> None:
    existed = store.get_node(complaint_id) is not None
    store.upsert_node(
        complaint_id,
        ["Complaint"],
        description=_string(description) or f"Complaint {complaint_id}",
        status="Closed" if "closed" in str(properties.get("closure_status", "")).lower() else "Open",
        severity="High" if any(token in str(description).lower() for token in ["alarm", "display", "shutdown"]) else "Medium",
        date="2026-06-30",
        event_date="2026-06-30",
        **{key: _string(value) for key, value in properties.items() if value not in (None, "")},
    )
    store.upsert_edge(complaint_id, device_id, "REPORTED_ON", rationale="DocPlus+ postmarket complaint trace")
    if not existed:
        summary.complaints += 1


def _link_requirement_risk(store: JsonGraphStore, req_id: str, risk_id: str, rationale: str) -> None:
    store.upsert_edge(req_id, risk_id, "MITIGATES", rationale=rationale)
    store.upsert_edge(req_id, risk_id, "MITIGATED_BY", rationale=f"{rationale} (compatibility edge)")


def _xlsx_rows(path: Path | None) -> Iterable[dict[str, Any]]:
    if path is None or not path.exists():
        return []
    try:
        import openpyxl
    except Exception:
        return []

    try:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            iterator = sheet.iter_rows(values_only=True)
            try:
                headers = [_normalize_header(value) for value in next(iterator)]
            except StopIteration:
                return []

            rows: list[dict[str, Any]] = []
            for values in iterator:
                row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
                if any(value not in (None, "") for value in row.values()):
                    rows.append(row)
            return rows
        finally:
            workbook.close()
    except Exception:
        return []


def _artifact_metadata(files: dict[str, Path], root: Path) -> dict[str, dict[str, Any]]:
    metadata: dict[str, dict[str, Any]] = {}
    for row in _xlsx_rows(files.get("Source_Index.xlsx")):
        keys = [row.get("file_name"), row.get("relative_path")]
        for key in keys:
            if key:
                metadata[str(key).replace("\\", "/")] = dict(row)
    for row in _xlsx_rows(files.get("Document_Control_Register.xlsx")):
        keys = [row.get("relative_path")]
        file_name = Path(str(row.get("relative_path", ""))).name if row.get("relative_path") else None
        if file_name:
            keys.append(file_name)
        for key in keys:
            if key:
                metadata.setdefault(str(key).replace("\\", "/"), {}).update(dict(row))
    for path in files.values():
        try:
            rel_path = str(path.relative_to(root)).replace("\\", "/")
        except ValueError:
            continue
        metadata.setdefault(path.name, {}).setdefault("relative_path", rel_path)
        metadata.setdefault(rel_path, {}).setdefault("file_name", path.name)
    return metadata


def _metadata_for(path: Path, rel_path: str, metadata_by_key: dict[str, dict[str, Any]]) -> dict[str, Any]:
    metadata = {}
    metadata.update(metadata_by_key.get(path.name, {}))
    metadata.update(metadata_by_key.get(rel_path, {}))
    return metadata


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _clean_id(value: Any) -> str:
    text = _string(value)
    if not text:
        return ""
    return re.split(r"[,;|\s]+", text)[0].strip()


def _extract_requirement_ids(value: Any) -> list[str]:
    return sorted(set(re.findall(r"REQ-[A-Z]+-\d+", _string(value))))


def _string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _bool(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "y", "approved"}


def _objective_default(file_name: str, artifact_type: str) -> bool:
    lower = f"{file_name} {artifact_type}".lower()
    return any(term in lower for term in ["test_report", "rawlog", "verification_evidence", "release_notes"])


def _row_text(row: dict[str, Any]) -> str:
    return " ".join(_string(value) for value in row.values())


def _infer_firmware_from_text(value: Any) -> str:
    match = re.search(r"(?:FW[-_\s]*)?v?(\d+(?:\.\d+)+)", _string(value), flags=re.IGNORECASE)
    return f"v{match.group(1)}" if match else ""


def _artifact_type(file_name: str) -> str:
    lower = file_name.lower()
    if "rtm" in lower:
        return "TraceabilityMatrix"
    if "hazard" in lower or "risk" in lower:
        return "RiskManagement"
    if "req_" in lower or "srs" in lower:
        return "RequirementSpecification"
    if "test_report" in lower:
        return "VerificationReport"
    if "rawlog" in lower:
        return "TelemetryLog"
    if "release_notes" in lower:
        return "FirmwareReleaseNotes"
    return "M1SourceArtifact"


def _file_digest(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _evidence_id(stem: str) -> str:
    return f"EVID-{_slug(stem)}"


def _slug(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "-", value.upper()).strip("-")[:80]
